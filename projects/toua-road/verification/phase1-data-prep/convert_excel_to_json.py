#!/usr/bin/env python3
"""
東亜検証_0203.xlsx を Azure AI Search用のJSONに変換するスクリプト

出力:
  - data/projects.json: 工事サマリ（61件）
  - data/direct_costs.json: 直接工事費明細
  - data/indirect_costs.json: 間接費明細
"""

import json
import re
from pathlib import Path
import openpyxl

# 入力ファイル
EXCEL_PATH = "/Users/kawagoekeita/Documents/Agent/★東亜PJ/東亜検証_0203.xlsx"

# 出力ディレクトリ
OUTPUT_DIR = Path("/Users/kawagoekeita/Documents/Agent/★東亜PJ/data")

# Excelの列マッピング（A=1, B=2, ...）
COLUMNS = {
    # 列A-M: 工事情報サマリDB
    "folder": 1,              # A: 対象フォルダ
    "filename": 2,            # B: 対象資料
    "project_name": 3,        # C: 工事名
    "branch": 4,              # D: 支店
    "location": 5,            # E: 工事場所(都道府県)
    "work_days": 6,           # F: 実施工期
    "contract_amount": 7,     # G: 請負金額
    "contract_period": 8,     # H: 契約工期
    "file_url": 9,            # I: ファイルURL
    "file_name": 10,          # J: ファイル名
    "site_manager": 11,       # K: 現場代理人名
    "tech_manager": 12,       # L: 監理技術者名
    "project_number": 13,     # M: 工事番号

    # 列N-AA: 直接工事費明細表DB
    "level": 14,              # N: 階層
    "cost_code": 15,          # O: 原価工種コード
    "item_name": 16,          # P: 工種・種別・細別
    "specification": 17,      # Q: 規格
    "unit": 18,               # R: 単位
    "quantity": 19,           # S: 数量
    "unit_price": 20,         # T: 単価
    "amount": 21,             # U: 金額
    "per_quantity": 22,       # V: 当り数量
    "composition_rate": 23,   # W: 構成率
    "contractor": 24,         # X: 業者名
    "note": 25,               # Y: 摘要
    "user_code": 26,          # Z: ユーザーコード
    "remarks": 27,            # AA: 備考

    # 列AB-AF: 費目別内訳
    "material_cost": 28,      # AB: 材料費
    "labor_cost": 29,         # AC: 労務費
    "outsource_cost": 30,     # AD: 外注費
    "machine_cost": 31,       # AE: 機械費
    "transport_cost": 32,     # AF: 運搬費

    # 列AG-AJ: 間接費明細表DB
    "indirect_branch": 33,    # AG: 支店(間接費)
    "indirect_project": 34,   # AH: 工事名(間接費)
    "indirect_category": 35,  # AI: 区分
    "indirect_items": 36,     # AJ: 内訳項目
}


def safe_str(value):
    """値を安全に文字列に変換"""
    if value is None:
        return ""
    return str(value).strip()


def safe_number(value):
    """値を安全に数値に変換"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        # カンマ区切りの数値を変換
        cleaned = str(value).replace(",", "").strip()
        if cleaned == "" or cleaned == "-":
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def safe_int(value):
    """値を安全に整数に変換"""
    num = safe_number(value)
    if num is None:
        return None
    return int(num)


def generate_project_id(index: int) -> str:
    """工事IDを生成（連番）"""
    return f"project_{index:04d}"


def generate_direct_cost_id(index: int) -> str:
    """直接工事費明細IDを生成（連番）"""
    return f"direct_{index:06d}"


def generate_indirect_cost_id(index: int) -> str:
    """間接費明細IDを生成（連番）"""
    return f"indirect_{index:06d}"


def parse_indirect_items(items_str: str) -> list:
    """
    間接費内訳項目（AJ列）をパースして構造化

    入力例: "｜｜重機運搬費 | 式 | 1 |  | "
    出力: [{"item_name": "重機運搬費", "unit": "式", "quantity": 1, "unit_price": None, "amount": None}]
    """
    if not items_str:
        return []

    results = []
    # 行ごとに分割（改行がある場合）
    lines = items_str.split('\n') if '\n' in items_str else [items_str]

    for line in lines:
        # 先頭の「｜」を除去
        cleaned = re.sub(r'^[｜|]+', '', line).strip()
        if not cleaned:
            continue

        # パイプで分割
        parts = [p.strip() for p in cleaned.split('|')]
        if len(parts) >= 1 and parts[0]:
            item = {
                "item_name": parts[0] if len(parts) > 0 else "",
                "unit": parts[1] if len(parts) > 1 and parts[1] else None,
                "quantity": safe_number(parts[2]) if len(parts) > 2 else None,
                "unit_price": safe_number(parts[3]) if len(parts) > 3 else None,
                "amount": safe_number(parts[4]) if len(parts) > 4 else None,
            }
            if item["item_name"]:  # 項目名がある場合のみ追加
                results.append(item)

    return results


def clean_item_name(name: str) -> str:
    """工種名から階層表示記号を除去"""
    if not name:
        return ""
    # 先頭の「｜」を除去
    return re.sub(r'^[｜|]+', '', name).strip()


def build_search_text(*fields) -> str:
    """検索用テキストを構築"""
    texts = [safe_str(f) for f in fields if f]
    return " ".join([t for t in texts if t])


def extract_data(excel_path: str):
    """Excelからデータを抽出（3ファイル用）"""
    print(f"読み込み中: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    projects = {}  # project_id -> project data
    direct_costs = []   # 直接工事費明細
    indirect_costs = []  # 間接費明細

    # 現在の工事情報（工事が切り替わるまで継承）
    current_project_info = None
    current_project_id = None
    project_index = 0
    direct_cost_index = 0
    indirect_cost_index = 0

    # 前の行のfolder/filename（変化検出用）
    prev_folder = None
    prev_filename = None

    # ヘッダー行をスキップ（1-2行目）、3行目からデータ開始
    for row_num in range(3, sheet.max_row + 1):
        # 工事情報の確認（列A-M）
        folder = safe_str(sheet.cell(row=row_num, column=COLUMNS["folder"]).value)
        filename = safe_str(sheet.cell(row=row_num, column=COLUMNS["filename"]).value)
        project_name = safe_str(sheet.cell(row=row_num, column=COLUMNS["project_name"]).value)

        # 新しい工事が始まる場合（folder/filenameの組み合わせが変わった時）
        is_new_project = (folder != prev_folder or filename != prev_filename) and folder and filename
        prev_folder = folder
        prev_filename = filename

        if is_new_project:
            project_index += 1
            current_project_id = generate_project_id(project_index)
            current_project_info = {
                "folder": folder,
                "filename": filename,
                "project_name": project_name,
                "branch": safe_str(sheet.cell(row=row_num, column=COLUMNS["branch"]).value),
                "location": safe_str(sheet.cell(row=row_num, column=COLUMNS["location"]).value),
                "work_days": safe_str(sheet.cell(row=row_num, column=COLUMNS["work_days"]).value),
                "contract_amount": safe_int(sheet.cell(row=row_num, column=COLUMNS["contract_amount"]).value),
                "contract_period": safe_str(sheet.cell(row=row_num, column=COLUMNS["contract_period"]).value),
                "file_url": safe_str(sheet.cell(row=row_num, column=COLUMNS["file_url"]).value),
                "file_name": safe_str(sheet.cell(row=row_num, column=COLUMNS["file_name"]).value),
                "site_manager": safe_str(sheet.cell(row=row_num, column=COLUMNS["site_manager"]).value),
                "tech_manager": safe_str(sheet.cell(row=row_num, column=COLUMNS["tech_manager"]).value),
                "project_number": safe_str(sheet.cell(row=row_num, column=COLUMNS["project_number"]).value),
            }

        # 工事情報がまだない場合はスキップ
        if current_project_info is None:
            continue

        # === 直接工事費明細（列N-AF） ===
        level = safe_int(sheet.cell(row=row_num, column=COLUMNS["level"]).value)
        item_name_raw = safe_str(sheet.cell(row=row_num, column=COLUMNS["item_name"]).value)
        item_name = clean_item_name(item_name_raw)
        specification = safe_str(sheet.cell(row=row_num, column=COLUMNS["specification"]).value)
        note = safe_str(sheet.cell(row=row_num, column=COLUMNS["note"]).value)
        remarks = safe_str(sheet.cell(row=row_num, column=COLUMNS["remarks"]).value)

        # 直接工事費の行チェック（levelまたはitem_nameがある行）
        if level is not None or item_name:
            direct_cost_index += 1
            direct_cost = {
                "id": generate_direct_cost_id(direct_cost_index),
                "project_id": current_project_id,
                # 工事情報（非正規化）
                "folder": current_project_info["folder"],
                "filename": current_project_info["filename"],
                "project_name": current_project_info["project_name"],
                "branch": current_project_info["branch"],
                "location": current_project_info["location"],
                "work_days": current_project_info["work_days"],
                "contract_amount": current_project_info["contract_amount"],
                "contract_period": current_project_info["contract_period"],
                "file_url": current_project_info["file_url"],
                "file_name": current_project_info["file_name"],
                "site_manager": current_project_info["site_manager"],
                "tech_manager": current_project_info["tech_manager"],
                "project_number": current_project_info["project_number"],
                # 直接工事費明細（列N-AF）
                "level": level,
                "cost_code": safe_str(sheet.cell(row=row_num, column=COLUMNS["cost_code"]).value),
                "item_name": item_name,
                "specification": specification,
                "unit": safe_str(sheet.cell(row=row_num, column=COLUMNS["unit"]).value),
                "quantity": safe_number(sheet.cell(row=row_num, column=COLUMNS["quantity"]).value),
                "unit_price": safe_number(sheet.cell(row=row_num, column=COLUMNS["unit_price"]).value),
                "amount": safe_number(sheet.cell(row=row_num, column=COLUMNS["amount"]).value),
                "per_quantity": safe_number(sheet.cell(row=row_num, column=COLUMNS["per_quantity"]).value),
                "composition_rate": safe_number(sheet.cell(row=row_num, column=COLUMNS["composition_rate"]).value),
                "contractor": safe_str(sheet.cell(row=row_num, column=COLUMNS["contractor"]).value),
                "note": note,
                "user_code": safe_str(sheet.cell(row=row_num, column=COLUMNS["user_code"]).value),
                "remarks": remarks,
                # 費目別内訳（列AB-AF）
                "material_cost": safe_number(sheet.cell(row=row_num, column=COLUMNS["material_cost"]).value),
                "labor_cost": safe_number(sheet.cell(row=row_num, column=COLUMNS["labor_cost"]).value),
                "outsource_cost": safe_number(sheet.cell(row=row_num, column=COLUMNS["outsource_cost"]).value),
                "machine_cost": safe_number(sheet.cell(row=row_num, column=COLUMNS["machine_cost"]).value),
                "transport_cost": safe_number(sheet.cell(row=row_num, column=COLUMNS["transport_cost"]).value),
                # 検索用テキスト
                "search_text": build_search_text(item_name, specification, note, remarks),
            }
            direct_costs.append(direct_cost)

            # 工事サマリを集約
            if current_project_id not in projects:
                projects[current_project_id] = {
                    "id": current_project_id,
                    **current_project_info,
                    "item_keywords": set(),
                    "total_items": 0,
                    "total_amount": 0,
                }

            # 工種名をキーワードに追加（階層3以下の意味のある工種名）
            if item_name and level is not None and level >= 3:
                projects[current_project_id]["item_keywords"].add(item_name)

            projects[current_project_id]["total_items"] += 1
            if direct_cost["amount"]:
                projects[current_project_id]["total_amount"] += direct_cost["amount"]

        # === 間接費明細（列AG-AJ） ===
        indirect_branch = safe_str(sheet.cell(row=row_num, column=COLUMNS["indirect_branch"]).value)
        indirect_project = safe_str(sheet.cell(row=row_num, column=COLUMNS["indirect_project"]).value)
        indirect_category = safe_str(sheet.cell(row=row_num, column=COLUMNS["indirect_category"]).value)
        indirect_items_raw = safe_str(sheet.cell(row=row_num, column=COLUMNS["indirect_items"]).value)

        # 間接費の行チェック（categoryまたはitemsがある行）
        if indirect_category or indirect_items_raw:
            # AJ列をパースして内訳項目を抽出
            parsed_items = parse_indirect_items(indirect_items_raw)

            if parsed_items:
                for item in parsed_items:
                    indirect_cost_index += 1
                    indirect_cost = {
                        "id": generate_indirect_cost_id(indirect_cost_index),
                        "project_id": current_project_id,
                        "project_name": current_project_info["project_name"],
                        "branch": indirect_branch or current_project_info["branch"],
                        "category": indirect_category,  # 共通仮設費、現場経費など
                        "item_name": item["item_name"],
                        "unit": item["unit"],
                        "quantity": item["quantity"],
                        "unit_price": item["unit_price"],
                        "amount": item["amount"],
                        "search_text": build_search_text(indirect_category, item["item_name"]),
                    }
                    indirect_costs.append(indirect_cost)
            elif indirect_category:
                # パースできなかったがcategoryがある場合
                indirect_cost_index += 1
                indirect_cost = {
                    "id": generate_indirect_cost_id(indirect_cost_index),
                    "project_id": current_project_id,
                    "project_name": current_project_info["project_name"],
                    "branch": indirect_branch or current_project_info["branch"],
                    "category": indirect_category,
                    "item_name": indirect_items_raw,
                    "unit": None,
                    "quantity": None,
                    "unit_price": None,
                    "amount": None,
                    "search_text": build_search_text(indirect_category, indirect_items_raw),
                }
                indirect_costs.append(indirect_cost)

        if row_num % 2000 == 0:
            print(f"処理中: {row_num}行")

    wb.close()

    # 工事サマリを整形
    project_list = []
    for project_id, project in projects.items():
        project["item_keywords"] = list(project["item_keywords"])
        project["search_text"] = build_search_text(
            project["project_name"],
            project["branch"],
            project["location"]
        )
        project_list.append(project)

    return project_list, direct_costs, indirect_costs


def main():
    # 出力ディレクトリ作成
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # データ抽出
    projects, direct_costs, indirect_costs = extract_data(EXCEL_PATH)

    print(f"\n抽出完了:")
    print(f"  - 工事数: {len(projects)}件")
    print(f"  - 直接工事費明細数: {len(direct_costs)}件")
    print(f"  - 間接費明細数: {len(indirect_costs)}件")

    # JSON出力
    projects_path = OUTPUT_DIR / "projects.json"
    direct_costs_path = OUTPUT_DIR / "direct_costs.json"
    indirect_costs_path = OUTPUT_DIR / "indirect_costs.json"

    with open(projects_path, "w", encoding="utf-8") as f:
        json.dump(projects, f, ensure_ascii=False, indent=2)
    print(f"\n出力: {projects_path}")

    with open(direct_costs_path, "w", encoding="utf-8") as f:
        json.dump(direct_costs, f, ensure_ascii=False, indent=2)
    print(f"出力: {direct_costs_path}")

    with open(indirect_costs_path, "w", encoding="utf-8") as f:
        json.dump(indirect_costs, f, ensure_ascii=False, indent=2)
    print(f"出力: {indirect_costs_path}")

    # サンプル表示
    print("\n=== 工事サンプル ===")
    if projects:
        sample = projects[0]
        print(json.dumps({
            "id": sample["id"],
            "project_name": sample["project_name"],
            "branch": sample["branch"],
            "total_items": sample["total_items"],
            "item_keywords_count": len(sample["item_keywords"]),
        }, ensure_ascii=False, indent=2))

    print("\n=== 直接工事費サンプル（階層3） ===")
    for d in direct_costs:
        if d["level"] == 3 and d["item_name"]:
            print(json.dumps({
                "id": d["id"],
                "project_id": d["project_id"],
                "project_name": d["project_name"],
                "level": d["level"],
                "item_name": d["item_name"],
                "specification": d["specification"],
                "quantity": d["quantity"],
                "unit_price": d["unit_price"],
                "amount": d["amount"],
            }, ensure_ascii=False, indent=2))
            break

    print("\n=== 間接費サンプル ===")
    for ic in indirect_costs[:3]:
        print(json.dumps({
            "id": ic["id"],
            "project_id": ic["project_id"],
            "project_name": ic["project_name"],
            "category": ic["category"],
            "item_name": ic["item_name"],
            "unit": ic["unit"],
            "quantity": ic["quantity"],
        }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
