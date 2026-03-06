"""
歩掛りデータETLスクリプト
Excel → JSON → Azure AI Search インデックス用

使い方:
  python sample_etl.py
"""

import openpyxl
import json
import os
from typing import List, Dict, Any

def extract_cost_items(sheet, start_row: int, parent_level: int) -> List[Dict[str, Any]]:
    """階層4の内訳データを抽出"""
    details = []
    row_num = start_row

    while row_num <= sheet.max_row:
        level = sheet.cell(row=row_num, column=1).value

        # 親レベル以下に戻ったら終了
        if level is not None and level <= parent_level:
            break

        # 階層4（内訳）のデータを抽出
        if level == 4:
            detail = {
                'cost_type': str(sheet.cell(row=row_num, column=3).value or ''),
                'item_name': str(sheet.cell(row=row_num, column=4).value or '').replace('｜', ''),
                'specification': str(sheet.cell(row=row_num, column=5).value or ''),
                'unit': str(sheet.cell(row=row_num, column=6).value or ''),
                'quantity': sheet.cell(row=row_num, column=7).value,
                'unit_price': sheet.cell(row=row_num, column=8).value,
                'amount': sheet.cell(row=row_num, column=9).value,
            }
            details.append(detail)

        row_num += 1

    return details, row_num

def build_search_text(item_name: str, specification: str, details: List[Dict]) -> str:
    """検索用テキストを構築（親＋内訳の情報を結合）"""
    texts = [item_name, specification]

    for detail in details:
        texts.append(detail.get('item_name', ''))
        texts.append(detail.get('specification', ''))

    # 空文字を除去して結合
    return ' '.join([t for t in texts if t])

def extract_project_data(filepath: str) -> List[Dict[str, Any]]:
    """Excelファイルから歩掛りデータを抽出"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['工事費一覧表(直接工事費)']

    file_name = os.path.basename(filepath)
    project_name = str(sheet.cell(row=1, column=4).value or '')

    documents = []
    current_category = ''  # 大分類（階層2）

    row_num = 5  # データ開始行
    while row_num <= sheet.max_row:
        level = sheet.cell(row=row_num, column=1).value
        ledger_type = sheet.cell(row=row_num, column=3).value

        # 階層2: 大分類を記録
        if level == 2:
            current_category = str(sheet.cell(row=row_num, column=4).value or '').replace('｜', '')
            row_num += 1
            continue

        # 階層3 + 内訳代価: 歩掛りデータの親
        if level == 3 and ledger_type == '内訳代価':
            item_name = str(sheet.cell(row=row_num, column=4).value or '').replace('｜', '')
            specification = str(sheet.cell(row=row_num, column=5).value or '')

            # 内訳（階層4）を抽出
            details, next_row = extract_cost_items(sheet, row_num + 1, 3)

            # 検索用テキスト構築
            search_text = build_search_text(item_name, specification, details)

            # ドキュメント作成
            doc = {
                'id': f"{file_name.replace('.xlsm', '')}_{row_num}",
                'file_name': file_name,
                'project_name': project_name,
                'category': current_category,
                'item_name': item_name,
                'specification': specification,
                'search_text': search_text,
                'unit': str(sheet.cell(row=row_num, column=6).value or ''),
                'quantity': sheet.cell(row=row_num, column=7).value,
                'unit_price': sheet.cell(row=row_num, column=8).value,
                'amount': sheet.cell(row=row_num, column=9).value,
                'details': details
            }
            documents.append(doc)

            row_num = next_row
            continue

        row_num += 1

    wb.close()
    return documents

def main():
    # 対象ファイル
    excel_path = '202601_提供データ（金塚）/厚木/2320710263.xlsm'

    print(f'処理中: {excel_path}')
    documents = extract_project_data(excel_path)

    print(f'抽出件数: {len(documents)}件')

    # JSON出力
    output_path = 'index_documents.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(documents, f, ensure_ascii=False, indent=2)

    print(f'出力完了: {output_path}')

    # サンプル表示
    print('\n=== サンプルデータ（集水桝A） ===')
    for doc in documents:
        if '集水桝A' in doc['item_name']:
            print(json.dumps(doc, ensure_ascii=False, indent=2))
            break

    # 検索テストのシミュレーション
    print('\n=== 検索シミュレーション ===')
    query = '750 現場打 集水桝'
    print(f'クエリ: {query}')

    keywords = query.split()
    for doc in documents:
        # すべてのキーワードがsearch_textに含まれているか
        if all(kw in doc['search_text'] for kw in keywords):
            print(f'\nヒット: {doc["item_name"]}')
            print(f'  カテゴリ: {doc["category"]}')
            print(f'  規格: {doc["specification"]}')
            print(f'  金額: {doc["amount"]:,}円')
            print(f'  内訳件数: {len(doc["details"])}件')

if __name__ == '__main__':
    main()
