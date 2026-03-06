"""パターンA/B: 実行予算書鑑ベースのExcelファイルを解析

対象フォルダ: 202510_事前提供データ, 金塚/厚木, 金塚/横浜, 岩手, 宮城, 中部
シート構成: 17枚（宮城は16枚 = 調書(外注費)なし）
"""

from pathlib import Path
import openpyxl
from .base import (
    ParseResult, safe_str, safe_number, safe_int,
    clean_item_name, build_search_text, format_date_value,
)

BLOB_BASE_URL = "https://toadorofilestorage.blob.core.windows.net/toadoro-files"


def parse(file_path: str, project_index: int, direct_cost_start: int, indirect_cost_start: int) -> ParseResult:
    """1つのExcelファイルを解析してParseResultを返す"""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    folder_name = Path(file_path).parent.name
    filename = Path(file_path).name
    project_id = f"project_{project_index:04d}"
    blob_url = f"{BLOB_BASE_URL}/{folder_name}/{filename}"

    # === 実行予算書鑑シートからメタデータ取得 ===
    sheet = wb['実行予算書鑑']

    project_number = safe_str(sheet.cell(row=3, column=7).value)
    branch = safe_str(sheet.cell(row=6, column=7).value)
    project_name = safe_str(sheet.cell(row=10, column=9).value)
    contract_amount = safe_int(sheet.cell(row=14, column=9).value)

    # 施工場所: Row27, Col6（固定位置）
    location = safe_str(sheet.cell(row=27, column=6).value)

    # 契約工期: Row28, Col6（開始） ～ Col14（終了）
    cp_start = format_date_value(sheet.cell(row=28, column=6).value)
    cp_end = format_date_value(sheet.cell(row=28, column=14).value)
    contract_period = f"{cp_start} ～ {cp_end}" if cp_start or cp_end else ""

    # 実施工期: Row28, Col25（開始） ～ Col33（終了）
    wd_start = format_date_value(sheet.cell(row=28, column=25).value)
    wd_end = format_date_value(sheet.cell(row=28, column=33).value)
    work_days = f"{wd_start} ～ {wd_end}" if wd_start or wd_end else ""

    project = {
        "id": project_id,
        "folder": folder_name,
        "filename": filename,
        "project_name": project_name,
        "branch": branch,
        "location": location,
        "work_days": work_days,
        "contract_amount": contract_amount,
        "contract_period": contract_period,
        "file_url": blob_url,
        "file_name": filename,
        "site_manager": "",
        "tech_manager": "",
        "project_number": project_number,
        "item_keywords": [],
        "total_items": 0,
        "total_amount": 0,
        "search_text": build_search_text(project_name, branch, location),
    }

    # === 工事費一覧表(直接工事費)シートから明細を取得 ===
    direct_costs = []
    dc_index = direct_cost_start
    item_keywords = set()

    if '工事費一覧表(直接工事費)' in wb.sheetnames:
        sheet_dc = wb['工事費一覧表(直接工事費)']
        sort_order = 0

        for row in range(4, sheet_dc.max_row + 1):
            level = safe_int(sheet_dc.cell(row=row, column=1).value)
            cost_code = safe_str(sheet_dc.cell(row=row, column=2).value)
            ledger_type = safe_str(sheet_dc.cell(row=row, column=3).value)
            item_name_raw = safe_str(sheet_dc.cell(row=row, column=4).value)
            item_name = clean_item_name(item_name_raw)
            specification = safe_str(sheet_dc.cell(row=row, column=5).value)
            unit = safe_str(sheet_dc.cell(row=row, column=6).value)
            quantity = safe_number(sheet_dc.cell(row=row, column=7).value)
            unit_price = safe_number(sheet_dc.cell(row=row, column=8).value)
            amount = safe_number(sheet_dc.cell(row=row, column=9).value)
            per_quantity = safe_number(sheet_dc.cell(row=row, column=10).value)

            if not item_name and level is None:
                continue

            sort_order += 1
            dc_index += 1
            direct_costs.append({
                "id": f"direct_{dc_index:06d}",
                "project_id": project_id,
                "folder": folder_name,
                "filename": filename,
                "project_name": project_name,
                "branch": branch,
                "location": location,
                "work_days": work_days,
                "contract_amount": contract_amount,
                "contract_period": contract_period,
                "file_url": blob_url,
                "file_name": filename,
                "site_manager": "",
                "tech_manager": "",
                "project_number": project_number,
                "sort_order": sort_order,
                "level": level,
                "cost_code": cost_code,
                "ledger_type": ledger_type,
                "item_name": item_name,
                "specification": specification,
                "unit": unit,
                "quantity": quantity,
                "unit_price": unit_price,
                "amount": amount,
                "per_quantity": per_quantity,
                "composition_rate": None,
                "contractor": "",
                "note": "",
                "user_code": "",
                "remarks": "",
                "material_cost": None,
                "labor_cost": None,
                "outsource_cost": None,
                "machine_cost": None,
                "transport_cost": None,
                "search_text": build_search_text(item_name, specification),
            })

            if item_name and level is not None and level >= 3:
                item_keywords.add(item_name)

            if amount:
                project["total_amount"] += amount

    project["item_keywords"] = list(item_keywords)
    project["total_items"] = len(direct_costs)

    # === 工事費一覧表(共通仮設費) / 工事費一覧表(現場経費) から間接費を取得 ===
    indirect_costs = []
    ic_index = indirect_cost_start

    for sheet_name, category in [
        ('工事費一覧表(共通仮設費)', '共通仮設費'),
        ('工事費一覧表(現場経費)', '現場経費'),
    ]:
        if sheet_name not in wb.sheetnames:
            continue

        sheet_ic = wb[sheet_name]
        sort_order = 0

        for row in range(4, sheet_ic.max_row + 1):
            level = safe_int(sheet_ic.cell(row=row, column=1).value)
            cost_code = safe_str(sheet_ic.cell(row=row, column=2).value)
            ledger_type = safe_str(sheet_ic.cell(row=row, column=3).value)
            item_name_raw = safe_str(sheet_ic.cell(row=row, column=4).value)
            item_name = clean_item_name(item_name_raw)
            specification = safe_str(sheet_ic.cell(row=row, column=5).value)
            unit = safe_str(sheet_ic.cell(row=row, column=6).value)
            quantity = safe_number(sheet_ic.cell(row=row, column=7).value)
            unit_price = safe_number(sheet_ic.cell(row=row, column=8).value)
            amount = safe_number(sheet_ic.cell(row=row, column=9).value)
            per_quantity = safe_number(sheet_ic.cell(row=row, column=10).value)
            composition_rate = safe_number(sheet_ic.cell(row=row, column=12).value)
            contractor = safe_str(sheet_ic.cell(row=row, column=13).value)
            note = safe_str(sheet_ic.cell(row=row, column=14).value)

            if not item_name and level is None:
                continue

            sort_order += 1
            ic_index += 1
            indirect_costs.append({
                "id": f"indirect_{ic_index:06d}",
                "project_id": project_id,
                "folder": folder_name,
                "filename": filename,
                "project_name": project_name,
                "branch": branch,
                "location": location,
                "work_days": work_days,
                "contract_amount": contract_amount,
                "contract_period": contract_period,
                "file_url": blob_url,
                "file_name": filename,
                "site_manager": "",
                "tech_manager": "",
                "project_number": project_number,
                "category": category,
                "sort_order": sort_order,
                "level": level,
                "cost_code": cost_code,
                "ledger_type": ledger_type,
                "item_name": item_name,
                "specification": specification,
                "unit": unit,
                "quantity": quantity,
                "unit_price": unit_price,
                "amount": amount,
                "per_quantity": per_quantity,
                "composition_rate": composition_rate,
                "contractor": contractor,
                "note": note,
                "search_text": build_search_text(category, item_name, specification),
            })

    wb.close()

    return ParseResult(
        project=project,
        direct_costs=direct_costs,
        indirect_costs=indirect_costs,
        source_file=file_path,
        pattern="regional_budget",
    )
