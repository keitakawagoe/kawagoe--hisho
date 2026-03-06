"""パターンC: 桐山フォーマット（データ/内訳一覧シートのみ）

対象フォルダ: 202601_提供データ（桐山）
シート構成: 2枚（データ, 内訳一覧）
特徴: コスト明細のみ。実施工期・契約工期・施工場所などメタデータなし。
"""

from pathlib import Path
import openpyxl
from .base import (
    ParseResult, safe_str, safe_number, safe_int,
    clean_item_name, build_search_text,
)

BLOB_BASE_URL = "https://toadorofilestorage.blob.core.windows.net/toadoro-files"


def parse(file_path: str, project_index: int, direct_cost_start: int, indirect_cost_start: int) -> ParseResult:
    """1つのExcelファイルを解析してParseResultを返す"""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    folder_name = Path(file_path).parent.name
    filename = Path(file_path).name
    project_id = f"project_{project_index:04d}"
    blob_url = f"{BLOB_BASE_URL}/{folder_name}/{filename}"

    # === 内訳一覧シートからメタデータ取得 ===
    sheet = wb['内訳一覧']
    project_name = safe_str(sheet.cell(row=1, column=5).value)   # E1
    contract_amount = safe_int(sheet.cell(row=1, column=9).value)  # I1

    project = {
        "id": project_id,
        "folder": folder_name,
        "filename": filename,
        "project_name": project_name,
        "branch": "",
        "location": "",
        "work_days": "",
        "contract_amount": contract_amount,
        "contract_period": "",
        "file_url": blob_url,
        "file_name": filename,
        "site_manager": "",
        "tech_manager": "",
        "project_number": "",
        "item_keywords": [],
        "total_items": 0,
        "total_amount": 0,
        "search_text": build_search_text(project_name),
    }

    # === 内訳一覧シートから直接工事費を取得 ===
    direct_costs = []
    dc_index = direct_cost_start
    item_keywords = set()

    sort_order = 0
    for row in range(2, sheet.max_row + 1):
        level = safe_int(sheet.cell(row=row, column=1).value)        # A
        ledger_type = safe_str(sheet.cell(row=row, column=3).value)  # C
        cost_code = safe_str(sheet.cell(row=row, column=4).value)    # D
        item_name_raw = safe_str(sheet.cell(row=row, column=5).value)  # E
        item_name = clean_item_name(item_name_raw)
        specification = safe_str(sheet.cell(row=row, column=6).value)  # F
        unit = safe_str(sheet.cell(row=row, column=7).value)         # G
        quantity = safe_number(sheet.cell(row=row, column=8).value)  # H
        unit_price = safe_number(sheet.cell(row=row, column=9).value)  # I
        amount = safe_number(sheet.cell(row=row, column=10).value)   # J

        if not item_name and level is None:
            continue

        sort_order += 1
        dc_index += 1
        direct_costs.append({
            "id": f"direct_{dc_index:06d}",
            "project_id": project_id,
            "folder": folder_name,
            "filename": filename,
            "project_name": project_name,
            "branch": "",
            "location": "",
            "work_days": "",
            "contract_amount": contract_amount,
            "contract_period": "",
            "file_url": blob_url,
            "file_name": filename,
            "site_manager": "",
            "tech_manager": "",
            "project_number": "",
            "sort_order": sort_order,
            "level": level,
            "cost_code": cost_code,
            "ledger_type": ledger_type,
            "item_name": item_name,
            "specification": specification,
            "unit": unit,
            "quantity": quantity,
            "unit_price": unit_price,
            "amount": amount,
            "per_quantity": None,
            "composition_rate": None,
            "contractor": "",
            "note": "",
            "user_code": "",
            "remarks": "",
            "material_cost": None,
            "labor_cost": None,
            "outsource_cost": None,
            "machine_cost": None,
            "transport_cost": None,
            "search_text": build_search_text(item_name, specification),
        })

        if item_name and level is not None and level >= 3:
            item_keywords.add(item_name)

        if amount:
            project["total_amount"] += amount

    project["item_keywords"] = list(item_keywords)
    project["total_items"] = len(direct_costs)

    wb.close()

    return ParseResult(
        project=project,
        direct_costs=direct_costs,
        indirect_costs=[],
        source_file=file_path,
        pattern="kiriyama_simple",
    )
