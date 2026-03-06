#!/usr/bin/env python3
"""
岩手・宮城フォルダのxlsmファイルを解析してAI Search用JSONに変換

出力:
  - data/projects_new.json: 工事サマリ（追加分）
  - data/direct_costs_new.json: 直接工事費明細（追加分）
  - data/indirect_costs_new.json: 間接費明細（追加分）
"""

import json
import re
from pathlib import Path
import openpyxl

# 入力フォルダ
INPUT_FOLDERS = [
    "/Users/kawagoekeita/Documents/Agent/★東亜PJ/岩手",
    "/Users/kawagoekeita/Documents/Agent/★東亜PJ/宮城",
]

# 出力ディレクトリ
OUTPUT_DIR = Path("/Users/kawagoekeita/Documents/Agent/★東亜PJ/data")

# Blob Storage設定
BLOB_BASE_URL = "https://toadorofilestorage.blob.core.windows.net/toadoro-files"

# 既存データの最大ID（これより大きい番号から採番）
EXISTING_PROJECT_COUNT = 61
EXISTING_DIRECT_COST_COUNT = 11313
EXISTING_INDIRECT_COST_COUNT = 243


def safe_str(value):
    """値を安全に文字列に変換"""
    if value is None:
        return ""
    return str(value).strip()


def safe_number(value):
    """値を安全に数値に変換"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        cleaned = str(value).replace(",", "").strip()
        if cleaned == "" or cleaned == "-":
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def safe_int(value):
    """値を安全に整数に変換"""
    num = safe_number(value)
    if num is None:
        return None
    return int(num)


def clean_item_name(name: str) -> str:
    """工種名から階層表示記号を除去"""
    if not name:
        return ""
    return re.sub(r'^[｜|]+', '', name).strip()


def build_search_text(*fields) -> str:
    """検索用テキストを構築"""
    texts = [safe_str(f) for f in fields if f]
    return " ".join([t for t in texts if t])


def get_folder_name(file_path: str) -> str:
    """ファイルパスからフォルダ名を取得"""
    return Path(file_path).parent.name


def get_blob_url(folder_name: str, filename: str) -> str:
    """Blob Storage URLを生成"""
    return f"{BLOB_BASE_URL}/{folder_name}/{filename}"


def parse_xlsm_file(file_path: str, project_index: int, direct_cost_start: int, indirect_cost_start: int):
    """
    1つのxlsmファイルを解析

    Returns:
        project: 工事情報dict
        direct_costs: 直接工事費リスト
        indirect_costs: 間接費リスト
        direct_cost_count: 追加された直接工事費数
        indirect_cost_count: 追加された間接費数
    """
    print(f"  解析中: {Path(file_path).name}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    folder_name = get_folder_name(file_path)
    filename = Path(file_path).name
    project_id = f"project_{project_index:04d}"

    # === 実行予算書鑑シートから工事情報を取得 ===
    sheet_kakan = wb['実行予算書鑑']

    project_number = safe_str(sheet_kakan.cell(row=3, column=7).value)
    branch = safe_str(sheet_kakan.cell(row=6, column=7).value)
    project_name = safe_str(sheet_kakan.cell(row=10, column=9).value)
    contract_amount = safe_int(sheet_kakan.cell(row=14, column=9).value)

    # 工事場所は工事概要シートから取得を試みる
    location = ""
    if '工事概要' in wb.sheetnames:
        sheet_gaiyo = wb['工事概要']
        # 工事場所は通常Row 5あたり
        for row in range(1, 15):
            for col in range(1, 10):
                val = safe_str(sheet_gaiyo.cell(row=row, column=col).value)
                if '工事場所' in val or '施工場所' in val:
                    # 次のセルか右のセルに場所がある
                    location = safe_str(sheet_gaiyo.cell(row=row, column=col+1).value)
                    if not location:
                        location = safe_str(sheet_gaiyo.cell(row=row+1, column=col).value)
                    break

    blob_url = get_blob_url(folder_name, filename)

    project = {
        "id": project_id,
        "folder": folder_name,
        "filename": filename,
        "project_name": project_name,
        "branch": branch,
        "location": location,
        "work_days": "",
        "contract_amount": contract_amount,
        "contract_period": "",
        "file_url": blob_url,
        "file_name": filename,
        "site_manager": "",
        "tech_manager": "",
        "project_number": project_number,
        "item_keywords": [],
        "total_items": 0,
        "total_amount": 0,
        "search_text": build_search_text(project_name, branch, location),
    }

    # === 工事費一覧表(直接工事費)シートから明細を取得 ===
    direct_costs = []
    direct_cost_index = direct_cost_start
    item_keywords = set()

    if '工事費一覧表(直接工事費)' in wb.sheetnames:
        sheet_direct = wb['工事費一覧表(直接工事費)']

        # Row 4からデータ開始（Row 2-3はヘッダー）
        for row in range(4, sheet_direct.max_row + 1):
            level = safe_int(sheet_direct.cell(row=row, column=1).value)
            cost_code = safe_str(sheet_direct.cell(row=row, column=2).value)
            item_name_raw = safe_str(sheet_direct.cell(row=row, column=4).value)
            item_name = clean_item_name(item_name_raw)
            specification = safe_str(sheet_direct.cell(row=row, column=5).value)
            unit = safe_str(sheet_direct.cell(row=row, column=6).value)
            quantity = safe_number(sheet_direct.cell(row=row, column=7).value)
            unit_price = safe_number(sheet_direct.cell(row=row, column=8).value)
            amount = safe_number(sheet_direct.cell(row=row, column=9).value)
            per_quantity = safe_number(sheet_direct.cell(row=row, column=10).value)

            # 空行はスキップ
            if not item_name and level is None:
                continue

            direct_cost_index += 1
            direct_cost = {
                "id": f"direct_{direct_cost_index:06d}",
                "project_id": project_id,
                "folder": folder_name,
                "filename": filename,
                "project_name": project_name,
                "branch": branch,
                "location": location,
                "work_days": "",
                "contract_amount": contract_amount,
                "contract_period": "",
                "file_url": blob_url,
                "file_name": filename,
                "site_manager": "",
                "tech_manager": "",
                "project_number": project_number,
                "level": level,
                "cost_code": cost_code,
                "item_name": item_name,
                "specification": specification,
                "unit": unit,
                "quantity": quantity,
                "unit_price": unit_price,
                "amount": amount,
                "per_quantity": per_quantity,
                "composition_rate": None,
                "contractor": "",
                "note": "",
                "user_code": "",
                "remarks": "",
                "material_cost": None,
                "labor_cost": None,
                "outsource_cost": None,
                "machine_cost": None,
                "transport_cost": None,
                "search_text": build_search_text(item_name, specification),
            }
            direct_costs.append(direct_cost)

            # キーワード収集（階層3以上）
            if item_name and level is not None and level >= 3:
                item_keywords.add(item_name)

            if amount:
                project["total_amount"] += amount

    project["item_keywords"] = list(item_keywords)
    project["total_items"] = len(direct_costs)

    # === 間接費一覧シートから間接費を取得 ===
    indirect_costs = []
    indirect_cost_index = indirect_cost_start

    if '間接費一覧' in wb.sheetnames:
        sheet_indirect = wb['間接費一覧']

        # 共通仮設費（Col 3: 項目名, Col 4: 金額）
        for row in range(4, 12):
            item_name = safe_str(sheet_indirect.cell(row=row, column=3).value)
            amount = safe_number(sheet_indirect.cell(row=row, column=4).value)

            if item_name and item_name != '計' and amount:
                indirect_cost_index += 1
                indirect_costs.append({
                    "id": f"indirect_{indirect_cost_index:06d}",
                    "project_id": project_id,
                    "project_name": project_name,
                    "branch": branch,
                    "category": "共通仮設費",
                    "item_name": item_name,
                    "unit": "式",
                    "quantity": 1.0,
                    "unit_price": amount,
                    "amount": amount,
                    "search_text": build_search_text("共通仮設費", item_name),
                })

        # 現場経費（Col 6: 項目名, Col 7: 金額）
        for row in range(4, 14):
            item_name = safe_str(sheet_indirect.cell(row=row, column=6).value)
            amount = safe_number(sheet_indirect.cell(row=row, column=7).value)

            if item_name and item_name != '計' and amount:
                indirect_cost_index += 1
                indirect_costs.append({
                    "id": f"indirect_{indirect_cost_index:06d}",
                    "project_id": project_id,
                    "project_name": project_name,
                    "branch": branch,
                    "category": "現場経費",
                    "item_name": item_name,
                    "unit": "式",
                    "quantity": 1.0,
                    "unit_price": amount,
                    "amount": amount,
                    "search_text": build_search_text("現場経費", item_name),
                })

    wb.close()

    return (
        project,
        direct_costs,
        indirect_costs,
        len(direct_costs),
        len(indirect_costs)
    )


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    all_projects = []
    all_direct_costs = []
    all_indirect_costs = []

    project_index = EXISTING_PROJECT_COUNT
    direct_cost_index = EXISTING_DIRECT_COST_COUNT
    indirect_cost_index = EXISTING_INDIRECT_COST_COUNT

    # 各フォルダを処理
    for folder_path in INPUT_FOLDERS:
        folder = Path(folder_path)
        if not folder.exists():
            print(f"フォルダが見つかりません: {folder_path}")
            continue

        print(f"\n=== {folder.name} フォルダ処理中 ===")

        xlsm_files = list(folder.glob("*.xlsm"))
        print(f"ファイル数: {len(xlsm_files)}")

        for xlsm_file in sorted(xlsm_files):
            try:
                project_index += 1
                project, direct_costs, indirect_costs, dc_count, ic_count = parse_xlsm_file(
                    str(xlsm_file),
                    project_index,
                    direct_cost_index,
                    indirect_cost_index
                )

                all_projects.append(project)
                all_direct_costs.extend(direct_costs)
                all_indirect_costs.extend(indirect_costs)

                direct_cost_index += dc_count
                indirect_cost_index += ic_count

            except Exception as e:
                print(f"    エラー: {e}")

    # 結果出力
    print(f"\n=== 解析完了 ===")
    print(f"  - 工事数: {len(all_projects)}件")
    print(f"  - 直接工事費明細数: {len(all_direct_costs)}件")
    print(f"  - 間接費明細数: {len(all_indirect_costs)}件")

    # JSON出力
    projects_path = OUTPUT_DIR / "projects_new.json"
    direct_costs_path = OUTPUT_DIR / "direct_costs_new.json"
    indirect_costs_path = OUTPUT_DIR / "indirect_costs_new.json"

    with open(projects_path, "w", encoding="utf-8") as f:
        json.dump(all_projects, f, ensure_ascii=False, indent=2)
    print(f"\n出力: {projects_path}")

    with open(direct_costs_path, "w", encoding="utf-8") as f:
        json.dump(all_direct_costs, f, ensure_ascii=False, indent=2)
    print(f"出力: {direct_costs_path}")

    with open(indirect_costs_path, "w", encoding="utf-8") as f:
        json.dump(all_indirect_costs, f, ensure_ascii=False, indent=2)
    print(f"出力: {indirect_costs_path}")

    # サンプル表示
    if all_projects:
        print("\n=== 工事サンプル ===")
        sample = all_projects[0]
        print(json.dumps({
            "id": sample["id"],
            "project_name": sample["project_name"],
            "branch": sample["branch"],
            "contract_amount": sample["contract_amount"],
            "total_items": sample["total_items"],
        }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
